import os
import openpyxl


def escape_sequence(message):
    print(message)
    exit()


def fetch_file():
    if len(os.listdir("./input/")) > 1:
        escape_sequence("more than one file in input directory")
    fn = os.listdir("./input/")[0]
    f_type = fn.split(".")[1]
    if f_type == "xlsx":
        f = openpyxl.load_workbook("./input/"+fn)
    else:
        escape_sequence("unsupported file type: " + f_type +". Unfortunately, we only support xlsx right now")
    return f_type, f


def get_row_num(wb):
    '''
    threshold indicates the number of cells allowed to be None before limiting the table size
    '''
    sheet = wb.active
    threshold = 5
    rows = []
    cols = 0
    nonecount_y = 0
    colnum = 0
    x = 1
    y = 0

    while nonecount_y <= threshold:
        rownum = 0
        y += 1
        # new row, so reinitialise x at 0
        nonecount_x = 0
        x = 1
        if sheet.cell(row = y, column = x).value is None:
            nonecount_y += 1
        else:
            nonecount_y = 0

        while nonecount_x <= threshold:
            cell_obj = sheet.cell(row = y, column = x)
            if cell_obj.value is None:
                nonecount_x += 1
            else:
                nonecount_x = 0
            rownum +=1
            x += 1
        rows.append(rownum) 
    # black magic
    colnum = y - threshold - 1
    rownum = max(rows) - threshold - 1
    return(colnum, rownum)


def make_blank_table(colnum, rownum):
    return([[""]*rownum]*colnum)


def fill_blank_table(blank, wb):
    sheet = wb.active
    fill = []
    for x, col in enumerate(blank):
        fill_row = []
        for y, row in enumerate(col):
            fill_row.append(str(sheet.cell(row = x+1, column = y+1).value or "").replace("\n", " "))
        fill.append(fill_row)
    return(fill)

def format_table(tb):
    '''
    input - nested array
    This function is to make the character length of every column the same - max char + 4
    output - nested array
    '''
    colchars = []
    for i in range(0, len(tb[0])):
        charlist = []
        colchars.append(max([len(c[i]) for c in tb]))

    formatted = []
    for row in tb:
        formatted_row = []
        for i, c in enumerate(row):
            formatted_row.append(c + " "*(colchars[i] - len(c) + 1))
        formatted.append(formatted_row)
    return(formatted)

def add_boundaries(tb):
    '''
    by default, we add a boundary around the whole table
    by default, we add boundaries between every column
    by default, we add a boundary above every row value that is not None
    '''
    # add side boundaries
    for row in tb:
        i = 0
        while i < len(tb[0]) + 1:
            if i%2:
                pass
            else:
                row.insert(i, "|")
            i += 1
   
    top_n_bottom = ["-"*len(c) for c in tb[0]]
    # add top bottom boundaries
    # assume that everything is grouped on index column (col[0])
    i = 0
    while i < len(tb):
        if len(tb[i][1].strip()):
            tb.insert(i, [t for t in top_n_bottom])
            i+=1
        i+=1
    tb.append(top_n_bottom)

    #finally, neaten up the joins
    for x, row in enumerate(tb):
        for y, col in enumerate(row):
            if col == "-":
                tb[x][y] = "+"
    print_table(tb)

def print_table(tb):
    for t in tb:
        print("".join(t))

def process_excel(wb):
    '''
    wb is a workbook object 
    first we construct boundaries - number of rows and number of cols
    '''
    colnum, rownum = get_row_num(wb)
    blank = make_blank_table(colnum, rownum)
    filled = fill_blank_table(blank, wb)
    formatted = format_table(filled)
    with_boundaries = add_boundaries(formatted)

def process_file(f, f_type):
    if f_type == "xlsx":
        process_excel(f)


def main():
    f_type, f = fetch_file()
    process_file(f, f_type)


if __name__=="__main__":
    main()
